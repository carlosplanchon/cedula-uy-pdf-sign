#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1"]
# ///
# Copyright 2026 Carlos Andrés Planchón Prestes
# Licensed under the Apache License, Version 2.0

"""Check every quotation in docs/security-invariants.md against the source that carries it.

A checker, never a generator, and the difference is the whole design. A generator that misparses
a source rewrites the page with a broken quotation and the authority of automation. A checker
that misparses fails red and a person looks. Both failure modes have already happened while this
page was being written: one fetch summarised a clause three words short, and one pass checked
against the catalogue alone and concluded that a sentence living in the guide did not exist. As
checker failures they were caught. As generator failures they would have shipped. So when this
script fails, the next step is a person reading what the source now says, never a script editing
the page.

Deliberately not in CI. It needs gub.uy and impo.com.uy to answer, and a test suite that goes
red because a third party is down is a suite that teaches people to ignore red. Run it by hand,
or from a scheduled job that opens an issue rather than blocking anyone:

    uv run scripts/check_mcu_quotes.py

Agesic publishes two artifacts and they are not interchangeable, which is the one fact this
script has to know. The requirement catalogue (the MCU 5.0 spreadsheets) gives each requirement
a title and the numbered text of every control, in a Cumplimiento sheet. The implementation
guide states an objective the catalogue does not carry and sorts the controls into levels the
catalogue does not show. The law is IMPO's. Each line of the page says which one it is quoting,
and this script checks it against exactly that one: a hit in the wrong artifact counts for
nothing.

It prints the SHA-256 of every artifact it fetched, so "checked" always means "checked against
these bytes" rather than against whatever the URLs serve next year.

Exit codes: 0 all quotations and level annotations verify, 1 at least one does not, 2 a source
could not be fetched or parsed, which says nothing about the page either way.
"""

from __future__ import annotations

import hashlib
import html
import io
import re
import sys
import urllib.error
import urllib.request
from pathlib import Path

import openpyxl

DOC = Path(__file__).resolve().parent.parent / "docs" / "security-invariants.md"

_GUB = ("https://www.gub.uy/agencia-gobierno-electronico-sociedad-informacion-conocimiento")
ARTIFACTS = {
    "catalogue-avanzado": _GUB + "/sites/agencia-gobierno-electronico-sociedad-informacion-"
                                 "conocimiento/files/documentos/publicaciones/"
                                 "Planilla%20MCU%205.0%20Avanzado.xlsx",
    "catalogue-basico": _GUB + "/sites/agencia-gobierno-electronico-sociedad-informacion-"
                               "conocimiento/files/documentos/publicaciones/"
                               "Planilla%20MCU%205.0%20B%C3%A1sico.xlsx",
    "catalogue-estandar": _GUB + "/sites/agencia-gobierno-electronico-sociedad-informacion-"
                                 "conocimiento/files/documentos/publicaciones/"
                                 "Planilla%20MCU%205.0%20Est%C3%A1ndar.xlsx",
    "guide-ad1": _GUB + "/comunicacion/publicaciones/guia-implementacion-del-mcu-50/"
                        "adquisicion-desarrollo-mantenimiento",
    "guide-ca4": _GUB + "/comunicacion/publicaciones/guia-implementacion-del-mcu-50/"
                        "control-acceso/ca4-establecer-controles",
    "law-art6": "https://www.impo.com.uy/bases/leyes/18600-2009/6",
}

# gub.uy answers curl and browsers and 403s bare library clients, so say something browser-like.
_HEADERS = {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64) firmauy-quote-check"}


def _fetch(url: str) -> bytes:
    request = urllib.request.Request(url, headers=_HEADERS)
    with urllib.request.urlopen(request, timeout=60) as response:
        return response.read()


def _text_of(raw: bytes) -> str:
    """HTML to comparable text. IMPO serves Latin-1 and gub.uy serves UTF-8, so try UTF-8
    first: Latin-1 never fails, and trying it first would silently mangle a UTF-8 page."""
    try:
        decoded = raw.decode("utf-8")
    except UnicodeDecodeError:
        decoded = raw.decode("latin-1")
    return _norm(html.unescape(re.sub(r"<[^>]+>", " ", decoded)))


def _norm(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def _sheets(raw: bytes) -> dict[str, str]:
    """Every sheet of a workbook as normalised text, keyed by sheet name."""
    workbook = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    return {
        sheet.title: _norm(" ".join(
            " ".join(str(cell) for cell in row if cell is not None)
            for row in sheet.iter_rows(values_only=True)))
        for sheet in workbook.worksheets
    }


def _ca4_controls(raw: bytes) -> dict[int, str]:
    """Every CA.4 control in the Cumplimiento sheet, by number, read cell by cell.

    Cell by cell and from that one sheet on purpose. The workbooks carry each control twice,
    once more under Madurez with different casing, and a regex over flattened sheet text once
    concluded from those boundaries that the tiers disagreed when they do not.
    """
    workbook = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    out = {}
    for sheet in workbook.worksheets:
        if sheet.title != "Cumplimiento":
            continue
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                match = cell and re.match(r"\s*CA\.4-(\d+):\s*(.+)", str(cell), re.S)
                if match:
                    out[int(match.group(1))] = _norm(match.group(2))
    return out


def _levels(guide_text: str, prefix: str) -> dict[int, int]:
    """Which level the guide sorts each numbered control into. The guide's Controles section is
    'Nivel N' headings with 'XX.N-M:' items under them, flattened here to text."""
    section = guide_text[guide_text.index("Controles"):]
    current, out = None, {}
    for match in re.finditer(rf"(Nivel (\d))|{prefix}-(\d+):", section):
        if match.group(1):
            current = int(match.group(2))
        else:
            out[int(match.group(3))] = current
    return out


def _entries(doc: str) -> list[tuple[str, int | None, str]]:
    """The verbatim list as (label, level annotation or None, quotation)."""
    section = doc[doc.index("## The clauses cited, verbatim"):doc.index("## One definition")]
    out = []
    for match in re.finditer(r"- \*\*\[([^\]]+)\]\*\*([^\"]*)\"([^\"]+)\"", section, re.S):
        level = re.search(r"\(level (\d)\)", match.group(2))
        out.append((match.group(1), int(level.group(1)) if level else None,
                    _norm(match.group(3))))
    return out


def main() -> int:
    doc = DOC.read_text()

    print("fetching:")
    raw = {}
    for name, url in ARTIFACTS.items():
        try:
            raw[name] = _fetch(url)
        except (urllib.error.URLError, TimeoutError) as exc:
            print(f"  {name}: UNREACHABLE ({exc}). Nothing verified either way. Try later.")
            return 2
        print(f"  {name}: {len(raw[name])} bytes  sha256 {hashlib.sha256(raw[name]).hexdigest()}")

    try:
        tiers = {name: _sheets(raw[name])
                 for name in ("catalogue-avanzado", "catalogue-basico", "catalogue-estandar")}
        guide_ad1 = _text_of(raw["guide-ad1"])
        guide_ca4 = _text_of(raw["guide-ca4"])
        law = _text_of(raw["law-art6"])
        levels = {"CA.4": _levels(guide_ca4, r"CA\.4"), "AD.1": _levels(guide_ad1, r"AD\.1")}
    except Exception as exc:  # a source changed shape, which is exactly worth a loud stop
        print(f"a source did not parse ({exc!r}): read it by hand before trusting anything")
        return 2

    catalogue_all = " ".join(" ".join(t.values()) for t in tiers.values())
    cumplimiento = {name: t.get("Cumplimiento", "") for name, t in tiers.items()}

    failures = 0

    def check(ok: bool, message: str) -> None:
        nonlocal failures
        failures += 0 if ok else 1
        print(f"  {'ok   ' if ok else 'FAIL '} {message}")

    print("\nstructural claims the page makes about the catalogue:")
    tiers_controls = {name: _ca4_controls(raw[name])
                      for name in ("catalogue-avanzado", "catalogue-basico", "catalogue-estandar")}
    counts = {name: len(found) for name, found in tiers_controls.items()}
    check(all(count == 12 for count in counts.values()),
          f"each tier carries twelve CA.4 controls in its Cumplimiento sheet ({counts})")
    reference = tiers_controls["catalogue-avanzado"]
    check(all(found == reference for found in tiers_controls.values()),
          "the twelve appear identically in all three tiers, as the page claims")

    print("\nquotations, each against the artifact its line declares:")
    for label, level, quote in _entries(doc):
        if label == "CA.4 title":
            check(quote in catalogue_all, f"[{label}] in the catalogue")
        elif label == "CA.4 objective":
            check(quote in guide_ca4, f"[{label}] in the guide")
            check(all(quote not in " ".join(t.values()) for t in tiers.values()),
                  "the objective stays absent from the catalogue, as the page claims")
        elif label.startswith("CA.4-"):
            # The page points auditors at the Cumplimiento sheet for these, so a hit anywhere
            # else in the workbook does not count. And it claims all three tiers agree.
            check(all(quote in sheet for sheet in cumplimiento.values()),
                  f"[{label}] in the Cumplimiento sheet of all three tiers")
        elif label.startswith("AD.1-"):
            check(quote in catalogue_all, f"[{label}] in the catalogue")
        elif label == "AD.1 evidencia":
            check(quote in guide_ad1, f"[{label}] in the guide")
        elif label.startswith("Ley"):
            check(quote in law, f"[{label}] in article 6 itself at IMPO")
        else:
            check(False, f"[{label}] declares no artifact this script knows how to check")

        if level is not None:
            family = label.split("-")[0]
            number = int(label.split("-")[1])
            check(levels.get(family, {}).get(number) == level,
                  f"[{label}] sits under level {level} in the guide")

    print(f"\n{failures} failing" if failures else "\neverything the page quotes, "
          "its source still says")
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
